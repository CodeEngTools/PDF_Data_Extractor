#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable, Optional
from decimal import Decimal, ROUND_HALF_UP

try:
    from pypdf import PdfReader
except ImportError:
    print("ERROR: falta dependencia. Instala con: pip install pypdf", file=sys.stderr)
    raise

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]


@dataclass
class InvoiceExtract:
    file: str
    invoice_no: Optional[str]
    pallets: Optional[int]
    boxes: Optional[int]
    nr_of_pack_pallets: Optional[int]  # si hay pallets usa pallets, si no boxes
    gross_weight: Optional[float]
    taxable_amount: Optional[float]
    total_invoice: Optional[float]


def reduce_repetition(token: str) -> str:
    """
    Reduce tokens tipo "...." o "----" repetidos a un único char.
    """
    if not token:
        return token
    return re.sub(r"([.\-_=])\1{2,}", r"\1", token)


def clean_text(s: str) -> str:
    s = s.replace("\x00", " ")
    s = s.replace("\u0000", " ")
    s = s.replace("\t", " ")
    s = s.replace("\r", "\n")
    s = re.sub(r"[ ]{2,}", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def iter_pdfs(inputs: list[Path]) -> Iterable[Path]:
    for p in inputs:
        if p.is_dir():
            yield from sorted(p.rglob("*.pdf"))
        else:
            if p.suffix.lower() == ".pdf":
                yield p


def extract_text_from_pdf(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        parts.append(t)
    return clean_text("\n".join(parts))


def m1(pattern: str, text: str, flags=re.IGNORECASE) -> Optional[str]:
    m = re.search(pattern, text, flags)
    if not m:
        return None
    return m.group(1).strip()


def parse_int(s: Optional[str]) -> Optional[int]:
    if not s:
        return None
    s = s.strip()
    s = s.replace(" ", "")
    # Integers: remove thousands separators (.)
    s = s.replace(".", "")
    s = s.replace(",", "")
    try:
        return int(s)
    except ValueError:
        return None


def parse_decimal_amount(s: Optional[str]) -> Optional[float]:
    """
    Importes tipo EUR:
    - Puede venir como 15.959,25 (EU) o 15959.25 (US)
    - Normalizamos a float con '.' decimal.
    """
    if not s:
        return None
    s = s.strip().replace(" ", "")

    if "," in s and "." in s:
        # Decide por el último separador como decimal
        if s.rfind(",") > s.rfind("."):
            # 15.959,25 -> 15959.25
            s = s.replace(".", "").replace(",", ".")
        else:
            # 15,959.25 -> 15959.25
            s = s.replace(",", "")
    elif "," in s:
        # 15959,25 -> 15959.25
        s = s.replace(".", "")  # por si acaso
        s = s.replace(",", ".")
    else:
        # 15959.25 o 15959
        s = s.replace(",", "")

    try:
        return float(s)
    except ValueError:
        return None


def parse_weight(s: Optional[str]) -> Optional[float]:
    """
    Pesos (net/gross) en estas facturas aparecen con decimales (p.ej. 592.828).
    Ojo: en otras facturas podría venir con separador de miles, pero aquí
    el ejemplo claro muestra decimal con '.'.
    - Si viene como 1.202.938 (miles), normalmente NO tiene 3 decimales "fijos"
      de peso, así que nos guiamos por el patrón:
        * si hay 1 punto y hay 3 dígitos después -> decimal (kg con 3 decimales)
        * si hay más de 1 punto -> miles -> eliminar puntos y tratar como entero
    """
    if not s:
        return None
    s = s.strip().replace(" ", "")

    dot_count = s.count(".")
    comma_count = s.count(",")

    # Caso con varios puntos: 1.202.938 -> 1202938
    if dot_count > 1 and comma_count == 0:
        s2 = s.replace(".", "")
        try:
            return float(s2)
        except ValueError:
            return None

    # Caso típico: 592.828 (3 decimales)
    if dot_count == 1 and comma_count == 0:
        left, right = s.split(".", 1)
        if len(right) == 3 and left.isdigit() and right.isdigit():
            try:
                return float(s)
            except ValueError:
                return None
        # si no es 3 decimales, interpretamos como miles
        s2 = s.replace(".", "")
        try:
            return float(s2)
        except ValueError:
            return None

    # Si viene coma decimal
    if comma_count == 1 and dot_count == 0:
        s2 = s.replace(",", ".")
        try:
            return float(s2)
        except ValueError:
            return None

    # Mixto raro: decide por último separador
    if comma_count and dot_count:
        if s.rfind(",") > s.rfind("."):
            s2 = s.replace(".", "").replace(",", ".")
        else:
            s2 = s.replace(",", "")
        try:
            return float(s2)
        except ValueError:
            return None

    # Fallback
    s2 = s.replace(",", "").replace(".", "")
    try:
        return float(s2)
    except ValueError:
        return None


def extract_fields(text: str, file_name: str) -> InvoiceExtract:
    # Invoice number: DM + 6 dígitos (preferible)
    invoice_no = None
    m = re.search(r"\b(DM\d{6})\b", text)
    if m:
        invoice_no = m.group(1)
    else:
        invoice_no = m1(r"Invoice Number:\s*(DM\d+)", text)

    pallets = parse_int(m1(r"Palets:\s*([0-9\.,]+)", text))
    boxes = parse_int(m1(r"Boxes:\s*([0-9\.,]+)", text))

    nr_of_pack_pallets = pallets if pallets and pallets > 0 else boxes

    gross_weight = parse_weight(m1(r"Gros Weight:\s*([0-9\.,]+)", text))
    taxable_amount = parse_decimal_amount(m1(r"Taxable Amount\s+([0-9\.,]+)", text))
    total_invoice = parse_decimal_amount(m1(r"Total Invoices\s+([0-9\.,]+)", text))

    return InvoiceExtract(
        file=file_name,
        invoice_no=invoice_no,
        pallets=pallets,
        boxes=boxes,
        nr_of_pack_pallets=nr_of_pack_pallets,
        gross_weight=gross_weight,
        taxable_amount=taxable_amount,
        total_invoice=total_invoice,
    )


def sum_decimal(values: Iterable[Optional[float]]) -> Decimal:
    total = Decimal("0")
    for v in values:
        if isinstance(v, (int, float)):
            # str(v) evita artefactos binarios típicos de float en el sumatorio
            total += Decimal(str(v))
    return total


def write_xlsx_from_rows(
    xlsx_path: Path,
    fieldnames: list[str],
    rows: list[InvoiceExtract],
    total_gross_weight: Decimal,
    total_total_invoice: Decimal,
) -> bool:
    """
    Genera un XLSX listo para Excel (sin avisos de locale/CSV), a partir de las filas ya extraídas.
    - Mantiene números como numéricos (no texto)
    - Formato: gross_weight con 4 decimales; importes con 2 decimales
    """
    if Workbook is None or get_column_letter is None:
        return False

    # Import local para evitar dependencia dura en el arranque si no está instalado.
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.worksheet import Worksheet

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "invoices"

    # Header
    ws.append(fieldnames)
    header_font = Font(bold=True)
    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    def _cell_value(col: str, v):
        if v is None:
            return None
        if col in ("pallets", "boxes", "nr_of_pack_pallets"):
            try:
                return int(v)
            except Exception:
                return None
        if col in ("gross_weight",):
            try:
                return float(v)
            except Exception:
                return None
        if col in ("taxable_amount", "total_invoice"):
            try:
                return float(v)
            except Exception:
                return None
        # strings (file, invoice_no)
        return str(v)

    for r in rows:
        d = asdict(r)
        ws.append([_cell_value(c, d.get(c)) for c in fieldnames])

    # Totals row (igual que en el CSV)
    ws.append(
        [
            _cell_value("file", "TOTAL"),
            None,
            None,
            None,
            None,
            float(total_gross_weight.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)),
            None,
            float(total_total_invoice.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
        ]
    )

    # Number formats
    col_map = {name: idx for idx, name in enumerate(fieldnames, start=1)}
    if "gross_weight" in col_map:
        gw_col = col_map["gross_weight"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=gw_col).number_format = "0.0000"
    for money_col in ("taxable_amount", "total_invoice"):
        if money_col in col_map:
            cidx = col_map[money_col]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=cidx).number_format = "0.00"

    # UX niceties
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{ws.max_row}"

    # Column widths (simple heuristic)
    for col_idx, name in enumerate(fieldnames, start=1):
        max_len = len(name)
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)

    wb.save(xlsx_path)
    return True


def main() -> int:
    ap = argparse.ArgumentParser(description="Extrae campos clave de PDFs con texto real (sin OCR).")
    ap.add_argument("inputs", nargs="+", type=Path, help="PDF(s) o carpeta(s) con PDFs")
    ap.add_argument("-o", "--out", type=Path, default=Path("out"), help="Carpeta de salida")
    args = ap.parse_args()

    out_dir: Path = args.out
    out_dir.mkdir(parents=True, exist_ok=True)

    rows: list[InvoiceExtract] = []
    for pdf in iter_pdfs(args.inputs):
        try:
            text = extract_text_from_pdf(pdf)
            rows.append(extract_fields(text, pdf.name))
        except Exception as e:
            print(f"[WARN] falló {pdf}: {e}", file=sys.stderr)

    # Totales
    total_gross_weight = sum_decimal(r.gross_weight for r in rows)
    total_total_invoice = sum_decimal(r.total_invoice for r in rows)

    total_gross_weight_r = total_gross_weight.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    total_total_invoice_r = total_total_invoice.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # JSON full
    json_path = out_dir / "invoices_extracted.json"
    json_path.write_text(
        json.dumps([asdict(r) for r in rows], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # JSON resumen
    summary = {
        "count": len(rows),
        "sum_gross_weight": float(total_gross_weight_r),
        "sum_total_invoice": float(total_total_invoice_r),
    }
    summary_path = out_dir / "invoices_extracted_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    # CSV + fila TOTAL
    csv_path = out_dir / "invoices_extracted.csv"
    fieldnames = [
        "file",
        "invoice_no",
        "pallets",
        "boxes",
        "nr_of_pack_pallets",
        "gross_weight",
        "taxable_amount",
        "total_invoice",
    ]

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(asdict(r))

        w.writerow(
            {
                "file": "TOTAL",
                "invoice_no": None,
                "pallets": None,
                "boxes": None,
                "nr_of_pack_pallets": None,
                "gross_weight": f"{total_gross_weight_r:.4f}",
                "taxable_amount": None,
                "total_invoice": f"{total_total_invoice_r:.2f}",
            }
        )

    # XLSX (para Excel)
    xlsx_path = out_dir / "invoices_extracted.xlsx"
    xlsx_ok = write_xlsx_from_rows(
        xlsx_path=xlsx_path,
        fieldnames=fieldnames,
        rows=rows,
        total_gross_weight=total_gross_weight,
        total_total_invoice=total_total_invoice,
    )

    print(f"OK → {json_path}")
    print(f"OK → {summary_path}")
    print(f"OK → {csv_path}")
    if xlsx_ok:
        print(f"OK → {xlsx_path}")
    else:
        print("WARN: openpyxl no está instalado; no se generó XLSX (instala: pip install openpyxl)", file=sys.stderr)

    print(f"SUM gross_weight: {total_gross_weight_r:.4f}")
    print(f"SUM total_invoice: {total_total_invoice_r:.2f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
