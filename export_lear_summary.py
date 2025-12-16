#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Iterable, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------
# Helpers: robust numeric casting
# -----------------------------

def _to_decimal(x: Any) -> Optional[Decimal]:
    if x is None:
        return None
    if isinstance(x, Decimal):
        return x
    if isinstance(x, (int, float)):
        # Convert via str to avoid float artifacts as much as possible
        return Decimal(str(x))
    s = str(x).strip()
    if not s:
        return None
    # remove NBSP and spaces
    s = s.replace("\u00A0", "").replace(" ", "")
    # allow only digits, separators, sign
    # NOTE: we assume lear_parser already normalized most values, but this makes exporter resilient.
    s2 = "".join(ch for ch in s if ch.isdigit() or ch in ",.-")
    if not s2:
        return None

    # If mixed ',' and '.', decide decimal by last separator (common robust rule)
    if "," in s2 and "." in s2:
        last_comma = s2.rfind(",")
        last_dot = s2.rfind(".")
        dec_sep = "," if last_comma > last_dot else "."
        thou_sep = "." if dec_sep == "," else ","
        s2 = s2.replace(thou_sep, "")
        s2 = s2.replace(dec_sep, ".")
    else:
        # If only comma, treat it as decimal separator
        if "," in s2 and "." not in s2:
            # if it looks like thousands (x,yyy with yyy 3 digits and short left), remove comma
            left, right = s2.split(",", 1)
            if right.isdigit() and len(right) == 3 and left.isdigit() and 1 <= len(left) <= 3:
                s2 = left + right
            else:
                s2 = left + "." + right

        # If only dot, keep it as decimal (parser should have normalized thousands already)

    try:
        return Decimal(s2)
    except (InvalidOperation, ValueError):
        return None


def _to_int(x: Any) -> Optional[int]:
    d = _to_decimal(x)
    if d is None:
        return None
    d = d.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return int(d)


def _to_money_2(x: Any) -> Optional[Decimal]:
    d = _to_decimal(x)
    if d is None:
        return None
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _safe_get(d: Dict[str, Any], path: str, default=None):
    """
    path example: "totals.total" or "extra.gross_weight"
    """
    cur: Any = d
    for part in path.split("."):
        if not isinstance(cur, dict):
            return default
        cur = cur.get(part)
        if cur is None:
            return default
    return cur


# -----------------------------
# Output schema
# -----------------------------

FIELDS = [
    "invoice_number",
    "issue_date",
    "currency",
    "subtotal",
    "tax",
    "total",
    "vendor_code",
    "customer_code",
    "number_of_pages",
    "number_of_lines",
    "parsed_lines_count",
    "taxable_amount_reported",
    "total_invoices_reported",   # (importe)
    "net_weight",                # entero
    "gross_weight",              # entero
]


@dataclass
class Row:
    invoice_number: str
    issue_date: str
    currency: str
    subtotal: Optional[Decimal]
    tax: Optional[Decimal]
    total: Optional[Decimal]
    vendor_code: str
    customer_code: str
    number_of_pages: Optional[int]
    number_of_lines: Optional[int]
    parsed_lines_count: Optional[int]
    taxable_amount_reported: Optional[Decimal]
    total_invoices_reported: Optional[Decimal]
    net_weight: Optional[int]
    gross_weight: Optional[int]

    def to_csv_dict(self) -> Dict[str, str]:
        # CSV as text; keep it stable. We do NOT add thousand separators.
        def fmt_dec(d: Optional[Decimal]) -> str:
            if d is None:
                return ""
            # dot decimal for machine; Excel may re-interpret; XLSX is the real product.
            return f"{d:.2f}"

        def fmt_int(i: Optional[int]) -> str:
            return "" if i is None else str(i)

        return {
            "invoice_number": self.invoice_number,
            "issue_date": self.issue_date,
            "currency": self.currency,
            "subtotal": fmt_dec(self.subtotal),
            "tax": fmt_dec(self.tax),
            "total": fmt_dec(self.total),
            "vendor_code": self.vendor_code,
            "customer_code": self.customer_code,
            "number_of_pages": fmt_int(self.number_of_pages),
            "number_of_lines": fmt_int(self.number_of_lines),
            "parsed_lines_count": fmt_int(self.parsed_lines_count),
            "taxable_amount_reported": fmt_dec(self.taxable_amount_reported),
            "total_invoices_reported": fmt_dec(self.total_invoices_reported),
            "net_weight": fmt_int(self.net_weight),
            "gross_weight": fmt_int(self.gross_weight),
        }


# -----------------------------
# Load + transform
# -----------------------------

def load_invoices(json_path: Path) -> List[Dict[str, Any]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        # common keys
        for k in ("invoices", "items", "documents"):
            v = data.get(k)
            if isinstance(v, list):
                return v
    raise ValueError(f"Unexpected JSON structure in {json_path}")


def build_rows(invoices: List[Dict[str, Any]]) -> List[Row]:
    rows: List[Row] = []
    for inv in invoices:
        extra = inv.get("extra") if isinstance(inv.get("extra"), dict) else {}
        totals = inv.get("totals") if isinstance(inv.get("totals"), dict) else {}

        invoice_number = str(inv.get("invoice_number") or "UNKNOWN")
        issue_date = str(inv.get("issue_date") or "UNKNOWN")

        currency = str(totals.get("currency") or extra.get("currency_reported") or "EUR")

        subtotal = _to_money_2(totals.get("subtotal"))
        tax = _to_money_2(totals.get("tax"))
        total = _to_money_2(totals.get("total"))

        vendor_code = str(extra.get("vendor_code") or "")
        customer_code = str(extra.get("customer_code") or "")

        number_of_pages = _to_int(extra.get("number_of_pages"))
        number_of_lines = _to_int(extra.get("number_of_lines"))
        parsed_lines_count = _to_int(extra.get("parsed_lines_count"))

        taxable_amount_reported = _to_money_2(extra.get("taxable_amount_reported"))
        total_invoices_reported = _to_money_2(extra.get("total_invoices_reported"))

        # Your rule: net/gross are integers ALWAYS
        net_weight = _to_int(extra.get("net_weight"))
        gross_weight = _to_int(extra.get("gross_weight"))

        rows.append(
            Row(
                invoice_number=invoice_number,
                issue_date=issue_date,
                currency=currency,
                subtotal=subtotal,
                tax=tax,
                total=total,
                vendor_code=vendor_code,
                customer_code=customer_code,
                number_of_pages=number_of_pages,
                number_of_lines=number_of_lines,
                parsed_lines_count=parsed_lines_count,
                taxable_amount_reported=taxable_amount_reported,
                total_invoices_reported=total_invoices_reported,
                net_weight=net_weight,
                gross_weight=gross_weight,
            )
        )
    return rows


# -----------------------------
# Write CSV (optional)
# -----------------------------

def write_csv(rows: List[Row], out_csv: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8-sig") as f:
        # Excel hint (Spain usually uses ;)
        f.write("sep=;\n")
        writer = csv.DictWriter(f, fieldnames=FIELDS, delimiter=";")
        writer.writeheader()
        for r in rows:
            writer.writerow(r.to_csv_dict())


# -----------------------------
# Write XLSX (solid Excel)
# -----------------------------

def write_xlsx(rows: List[Row], out_xlsx: Path) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "lear_summary"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    # Write header
    ws.append(FIELDS)
    for col_idx, name in enumerate(FIELDS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.alignment = header_align

    # Column formats (no thousands)
    # - money: €0.00 (without grouping)
    # - ints: 0
    money_fmt = '"€"0.00'
    dec2_fmt = '0.00'
    int_fmt = '0'

    money_cols = {"subtotal", "tax", "total", "taxable_amount_reported", "total_invoices_reported"}
    int_cols = {"number_of_pages", "number_of_lines", "parsed_lines_count", "net_weight", "gross_weight"}

    col_index = {name: i for i, name in enumerate(FIELDS, start=1)}

    # Write rows
    for r in rows:
        ws.append([
            r.invoice_number,
            r.issue_date,
            r.currency,
            float(r.subtotal) if r.subtotal is not None else None,
            float(r.tax) if r.tax is not None else None,
            float(r.total) if r.total is not None else None,
            r.vendor_code,
            r.customer_code,
            r.number_of_pages,
            r.number_of_lines,
            r.parsed_lines_count,
            float(r.taxable_amount_reported) if r.taxable_amount_reported is not None else None,
            float(r.total_invoices_reported) if r.total_invoices_reported is not None else None,
            r.net_weight,
            r.gross_weight,
        ])

    # Apply number formats
    max_row = ws.max_row
    for name, idx in col_index.items():
        if name in money_cols:
            fmt = money_fmt
        elif name in int_cols:
            fmt = int_fmt
        else:
            fmt = None

        if fmt:
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=idx)
                if cell.value is not None and cell.value != "":
                    cell.number_format = fmt

    # For totals columns, if you prefer no € symbol on subtotal/tax/total, switch to dec2_fmt:
    # (Leave total_invoices_reported as money_fmt)
    for name in ("subtotal", "tax", "total", "taxable_amount_reported"):
        idx = col_index[name]
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if cell.value is not None and cell.value != "":
                cell.number_format = dec2_fmt

    # Freeze header + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}1"

    # Reasonable column widths
    widths = {
        "invoice_number": 16,
        "issue_date": 12,
        "currency": 10,
        "subtotal": 14,
        "tax": 10,
        "total": 14,
        "vendor_code": 18,
        "customer_code": 18,
        "number_of_pages": 16,
        "number_of_lines": 16,
        "parsed_lines_count": 18,
        "taxable_amount_reported": 20,
        "total_invoices_reported": 22,
        "net_weight": 12,
        "gross_weight": 12,
    }
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(col_index[name])].width = w

    wb.save(out_xlsx)


# -----------------------------
# CLI
# -----------------------------

def main() -> int:
    p = argparse.ArgumentParser(description="Export Lear summary to CSV and XLSX (Excel-solid).")
    p.add_argument("--in", dest="input_json", required=True, help="Path to merged JSON (e.g. out/lear_merged.json)")
    p.add_argument("--out-dir", default="out", help="Output directory (default: out)")
    p.add_argument("--base-name", default="lear", help="Base name for outputs (default: lear)")
    p.add_argument("--no-csv", action="store_true", help="Do not write CSV (only XLSX)")
    args = p.parse_args()

    in_path = Path(args.input_json)
    out_dir = Path(args.out_dir)
    base = args.base_name

    invoices = load_invoices(in_path)
    rows = build_rows(invoices)

    out_xlsx = out_dir / f"{base}_summary.xlsx"
    write_xlsx(rows, out_xlsx)

    if not args.no_csv:
        out_csv = out_dir / f"{base}_summary.csv"
        write_csv(rows, out_csv)

    print(f"XLSX → {out_xlsx}")
    if not args.no_csv:
        print(f"CSV  → {out_dir / f'{base}_summary.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
