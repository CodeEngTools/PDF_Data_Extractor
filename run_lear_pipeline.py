from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from app.pdf_extractor import extract_pdf_text
from app.normalizer import normalize_text
from app.parsers.lear_parser import LearInvoiceParser


# ----------------------------
# Number parsing / formatting
# ----------------------------

_NUM_CLEAN_RE = re.compile(r"[^\d,\.\-]+")


def _normalize_number_str(s: str) -> Optional[str]:
    """
    Normaliza un número que puede venir en formato US/EU y/o con separadores de miles.
    Devuelve string en formato "1234.56" (punto decimal, sin miles) o None.

    Heurística pensada para tus PDFs:
      - Pesos a veces con 4 decimales: 196.4845
      - A veces miles con '.': 74.058  -> 74058
      - A veces miles múltiples: 1.202.938 -> 1202938
      - Importes: típicamente 2 decimales
    """
    if s is None:
        return None
    s = s.strip()
    if not s:
        return None

    s = _NUM_CLEAN_RE.sub("", s.replace(" ", ""))
    if not s or s in ("-", ".", ","):
        return None

    # Ambos separadores: decide el último como decimal
    if "." in s and "," in s:
        if s.rfind(",") > s.rfind("."):
            # EU: '.' miles, ',' decimal
            s = s.replace(".", "").replace(",", ".")
        else:
            # US: ',' miles, '.' decimal
            s = s.replace(",", "")
        return s

    # Solo comas
    if "," in s and "." not in s:
        if s.count(",") > 1:
            # múltiples comas -> miles
            return s.replace(",", "")
        left, right = s.split(",", 1)
        if 1 <= len(right) <= 4:
            left = left.replace(".", "")
            return f"{left}.{right}"
        return (left + right).replace(".", "")

    # Solo puntos
    if "." in s and "," not in s:
        if s.count(".") > 1:
            # múltiples puntos -> miles
            return s.replace(".", "")
        left, right = s.split(".", 1)
        # 3 dígitos y parte entera corta => miles (74.058 -> 74058)
        if len(right) == 3 and len(left) <= 3:
            return left + right
        return s

    return s


def _to_decimal(x: Any) -> Optional[Decimal]:
    if x is None:
        return None
    if isinstance(x, Decimal):
        return x
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        try:
            return Decimal(str(x))
        except InvalidOperation:
            return None
    if isinstance(x, str):
        norm = _normalize_number_str(x)
        if not norm:
            return None
        try:
            return Decimal(norm)
        except InvalidOperation:
            return None
    return None


def _fmt_csv_decimal(x: Any, decimals: int, decimal_sep: str = ",") -> str:
    d = _to_decimal(x)
    if d is None:
        return ""
    q = Decimal("1").scaleb(-decimals)  # 10^-decimals
    d = d.quantize(q, rounding=ROUND_HALF_UP)
    s = format(d, "f")
    if decimal_sep != ".":
        s = s.replace(".", decimal_sep)
    return s


def _fmt_csv_int(x: Any) -> str:
    d = _to_decimal(x)
    if d is None:
        return ""
    d = d.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return str(int(d))


# ----------------------------
# Parsing
# ----------------------------

def parse_invoice_pdf(pdf_path: Union[str, Path]) -> Dict[str, Any]:
    pdf_path = Path(pdf_path)
    raw_text = extract_pdf_text(str(pdf_path))
    text = normalize_text(raw_text)

    parser = LearInvoiceParser()
    invoice = parser.parse(text)
    return asdict(invoice)


def parse_invoices_in_folder(folder_path: Union[str, Path]) -> List[Dict[str, Any]]:
    folder = Path(folder_path)
    if not folder.is_dir():
        raise NotADirectoryError(f"{folder} no es un directorio")

    results: List[Dict[str, Any]] = []
    for pdf in sorted(folder.glob("*.pdf")):
        try:
            results.append(parse_invoice_pdf(pdf))
        except Exception as e:
            print(f"[WARN] Error procesando {pdf.name}: {e}")
    return results


def build_summary(invoices: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary: List[Dict[str, Any]] = []
    for inv in invoices:
        extra = inv.get("extra") or {}
        summary.append(
            {
                "invoice_number": inv.get("invoice_number"),
                "issue_date": inv.get("issue_date"),
                "pallets": extra.get("pallets"),
                "net_weight": extra.get("net_weight"),
                "gross_weight": extra.get("gross_weight"),
                "total_invoices_reported": extra.get("total_invoices_reported"),
            }
        )
    return summary


# ----------------------------
# Outputs
# ----------------------------

def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def write_summary_csv(
    summary: List[Dict[str, Any]],
    out_path: Path,
    *,
    delimiter: str = ";",
    decimal_sep: str = ",",
) -> None:
    """
    CSV “Excel-friendly” (EU):
      - delimiter = ;
      - decimal = ,
      - sin separadores de miles
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "invoice_number",
        "issue_date",
        "pallets",
        "net_weight",
        "gross_weight",
        "total_invoices_reported",
    ]

    total_pallets = Decimal("0")
    total_net_w = Decimal("0")
    total_gross_w = Decimal("0")
    total_amount = Decimal("0")

    rows: List[Dict[str, str]] = []
    for row in summary:
        p = _to_decimal(row.get("pallets")) or Decimal("0")
        nw = _to_decimal(row.get("net_weight")) or Decimal("0")
        gw = _to_decimal(row.get("gross_weight")) or Decimal("0")
        amt = _to_decimal(row.get("total_invoices_reported")) or Decimal("0")

        total_pallets += p
        total_net_w += nw
        total_gross_w += gw
        total_amount += amt

        rows.append(
            {
                "invoice_number": row.get("invoice_number") or "",
                "issue_date": row.get("issue_date") or "",
                "pallets": _fmt_csv_int(row.get("pallets")),
                "net_weight": _fmt_csv_decimal(row.get("net_weight"), 4, decimal_sep=decimal_sep),
                "gross_weight": _fmt_csv_decimal(row.get("gross_weight"), 4, decimal_sep=decimal_sep),
                "total_invoices_reported": _fmt_csv_decimal(row.get("total_invoices_reported"), 2, decimal_sep=decimal_sep),
            }
        )

    rows.append(
        {
            "invoice_number": "TOTAL",
            "issue_date": "",
            "pallets": _fmt_csv_int(total_pallets),
            "net_weight": _fmt_csv_decimal(total_net_w, 4, decimal_sep=decimal_sep),
            "gross_weight": _fmt_csv_decimal(total_gross_w, 4, decimal_sep=decimal_sep),
            "total_invoices_reported": _fmt_csv_decimal(total_amount, 2, decimal_sep=decimal_sep),
        }
    )

    with out_path.open("w", newline="", encoding="utf-8") as f:
        f.write(f"sep={delimiter}\n")  # ayuda a Excel con el delimiter
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)


def write_summary_xlsx(summary: List[Dict[str, Any]], out_path: Path) -> None:
    """
    XLSX “locale-proof”:
      - números reales (no texto)
      - formatos fijos:
          pallets: 0
          net/gross_weight: 0.0000
          amount: 0.00
      - sin separadores de miles
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except Exception:
        print("[WARN] openpyxl no disponible; se omite XLSX.")
        return

    fieldnames = [
        "invoice_number",
        "issue_date",
        "pallets",
        "net_weight",
        "gross_weight",
        "total_invoices_reported",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    ws.append(fieldnames)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(fieldnames) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = header_align

    col_idx = {name: i + 1 for i, name in enumerate(fieldnames)}

    widths = {
        "invoice_number": 16,
        "issue_date": 12,
        "pallets": 10,
        "net_weight": 16,
        "gross_weight": 16,
        "total_invoices_reported": 22,
    }
    for name, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx[name])].width = w

    num_align = Alignment(horizontal="right")

    total_pallets = Decimal("0")
    total_net_w = Decimal("0")
    total_gross_w = Decimal("0")
    total_amount = Decimal("0")

    for row in summary:
        inv = row.get("invoice_number") or ""
        date = row.get("issue_date") or ""

        p = _to_decimal(row.get("pallets"))
        nw = _to_decimal(row.get("net_weight"))
        gw = _to_decimal(row.get("gross_weight"))
        amt = _to_decimal(row.get("total_invoices_reported"))

        total_pallets += (p or Decimal("0"))
        total_net_w += (nw or Decimal("0"))
        total_gross_w += (gw or Decimal("0"))
        total_amount += (amt or Decimal("0"))

        ws.append(
            [
                inv,
                date,
                int(p) if p is not None else None,
                float(nw) if nw is not None else None,
                float(gw) if gw is not None else None,
                float(amt) if amt is not None else None,
            ]
        )

        r = ws.max_row

        c = ws.cell(row=r, column=col_idx["pallets"])
        c.number_format = "0"
        c.alignment = num_align

        for k in ("net_weight", "gross_weight"):
            c = ws.cell(row=r, column=col_idx[k])
            c.number_format = "0.0000"
            c.alignment = num_align

        c = ws.cell(row=r, column=col_idx["total_invoices_reported"])
        c.number_format = "0.00"
        c.alignment = num_align

    ws.append(
        [
            "TOTAL",
            None,
            int(total_pallets),
            float(total_net_w.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)),
            float(total_gross_w.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)),
            float(total_amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
        ]
    )
    r = ws.max_row
    ws.cell(row=r, column=col_idx["pallets"]).number_format = "0"
    ws.cell(row=r, column=col_idx["net_weight"]).number_format = "0.0000"
    ws.cell(row=r, column=col_idx["gross_weight"]).number_format = "0.0000"
    ws.cell(row=r, column=col_idx["total_invoices_reported"]).number_format = "0.00"
    for k in ("pallets", "net_weight", "gross_weight", "total_invoices_reported"):
        ws.cell(row=r, column=col_idx[k]).alignment = num_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{ws.max_row}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_outputs(
    invoices: List[Dict[str, Any]],
    out_dir: Union[str, Path],
    base_name: str = "lear",
) -> None:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    merged_json = out_dir / f"{base_name}_merged.json"
    summary_json = out_dir / f"{base_name}_summary.json"
    summary_csv = out_dir / f"{base_name}_summary.csv"
    summary_xlsx = out_dir / f"{base_name}_summary.xlsx"

    summary = build_summary(invoices)

    write_json(merged_json, invoices)
    write_json(summary_json, summary)
    write_summary_csv(summary, summary_csv)
    write_summary_xlsx(summary, summary_xlsx)

    print(f"JSON completo   → {merged_json}")
    print(f"JSON resumen    → {summary_json}")
    print(f"CSV resumen     → {summary_csv}")
    if summary_xlsx.exists():
        print(f"XLSX resumen    → {summary_xlsx}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("input_folder", help="Carpeta con PDFs")
    p.add_argument("-o", "--output-dir", required=True, help="Carpeta de salida")
    p.add_argument("--base-name", default="lear", help="Prefijo de los ficheros de salida")
    args = p.parse_args()

    invoices = parse_invoices_in_folder(args.input_folder)
    print(f"Facturas procesadas: {len(invoices)}")
    write_outputs(invoices, args.output_dir, base_name=args.base_name)


if __name__ == "__main__":
    main()
